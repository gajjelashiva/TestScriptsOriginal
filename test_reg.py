
import os
import xlrd
import glob

def readJavafile(javafilname):

    with open(javafilname) as f:
        s=f.read()
        import re
        #matches = re.findall("@WbxCase\(.*,title.?=(.*).description.=(.*)\)",s)
        matches = re.findall('@WbxCase\(.*,title.?=.?"(.*)"?,.description.=.?"(.*)\)',s)
        print(matches)
        for i,v in matches:
            print 'title is='+i+'description is'+v
        return matches
        
        
location='/Users/sgajjela/PycharmProjects/test/venv/test12.xlsx'
import openpyxl

# Give the location of the file


# to open the workbook
# workbook object is created
wb_obj = openpyxl.Workbook()
sheet_obj = wb_obj.active

def writetoxl(javafilname,worksheetname,index):
    ws1 = wb_obj.create_sheet(worksheetname,index)
    # print the total number of rows
    print(ws1.max_row, ws1.title)
    list_matche = readJavafile(javafilname)
    for i in range(0,len(list_matche)):
        for j in range(1,3):
            c1 = ws1.cell(row=i+1, column=j)
            c1.value = list_matche[i][j-1].strip('"')

path = '/Users/sgajjela/git/EVENT_CENTER/webex-web-eventcenter/uitest/src/main/java/com/cisco/webex/testcase/train/ec/schedule'
files = [f for f in glob.glob(path + "**/*.java")]
i=0
for file in files:
    writetoxl(file,file.split(os.sep)[-1].strip('.java'),i)
    i=i+1

wb_obj.save(location)